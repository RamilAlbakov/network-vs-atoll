"""Test fuctions from excel module."""

from network_vs_atoll.excel import fill_excel
from openpyxl import Workbook, load_workbook


def test_fill_excel(selected_data):
    """Test fill excel function."""
    report_path = 'tests/reports/network-vs-atoll.xlsx'
    fill_excel(selected_data, report_path)

    report_wb = load_workbook(report_path)
    lte_sheet = report_wb['lte']

    assert len(selected_data.keys()) == len(report_wb.sheetnames)
    assert lte_sheet.max_row == len(selected_data['lte']) + 1

