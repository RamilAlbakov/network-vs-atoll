"""Fill excel file with cells missing in atoll."""

from openpyxl import Workbook


def fill_sheet(sheet, missed_cells):
    """
    Fill the excel sheet with missed cells.

    Args:
        sheet: openpyxl sheet object
        missed_cells: list
    """
    columns = missed_cells[0]._fields

    row = 1
    for col in columns:
        sheet.cell(row=row, column=columns.index(col) + 1, value=col)

    row = 2
    for missed_cell in missed_cells:
        for cell_val in missed_cell:
            sheet.cell(row=row, column=missed_cell.index(cell_val) + 1, value=cell_val)
        row += 1


def fill_excel(selected_data, path):
    """
    Fill excel file with missed in atoll cells, each technology to separate sheet.

    Args:
        selected_data: dict
        path: string
    """
    work_book = Workbook()

    for tech, missed_cells in selected_data.items():
        work_sheet = work_book.create_sheet(tech)
        fill_sheet(work_sheet, missed_cells)

    del work_book['Sheet']
    work_book.save(path)
